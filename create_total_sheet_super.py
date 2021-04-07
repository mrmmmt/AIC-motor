# -*- coding: utf-8 -*-
import pandas as pd
from frozen_dir import app_path
import os
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta
from tkinter import Tk, messagebox, Label, StringVar, Button, ttk


def create_total_sheet_raw_df(file_name_0695, pkl_file_name, data_date):
    
    with open(pkl_file_name, 'rb') as f:
        pkl_file_df = pickle.load(f)
        
    pkl_file_df['投保日期'] = str(data_date)
    pkl_file_df[['保险起期', '保险止期', '订单时间_0366']] = pkl_file_df[['保险起期', '保险止期', '订单时间_0366']].apply(pd.to_datetime).astype('str').replace('/', '-')
    pkl_file_df['序号'] = range(pkl_file_df.shape[0])

    if file_name_0695.split('.')[-1] in ['xls', 'xlsx']:
        df_0695 = pd.read_excel(file_name_0695, dtype='str', header=1)
        df_0695 = df_0695.loc[:, ['保单号', '投保人证件号码', '导入时间']]
        df_0695 = df_0695.rename(columns={'保单号':'保单号_0695', '导入时间':'导入时间_0695'})
    else:
        df_0695 = pd.DataFrame(columns=['保单号_0695', '投保人证件号码', '导入时间_0695'])

    df_0695 = df_0695.drop_duplicates()

    pkl_file_df['temp'] = 1
    pkl_file_df['被保人身份证号_unique'] = pkl_file_df['被保人身份证号'] + '_' + pkl_file_df.groupby('被保人身份证号', as_index=False)['temp'].rank(method='first')['temp'].map(lambda x: str(int(x)).zfill(2))    
    df_0695['temp'] = 1
    df_0695['被保人身份证号_unique'] = df_0695['投保人证件号码'] + '_' + df_0695.groupby('投保人证件号码', as_index=False)['temp'].rank(method='first')['temp'].map(lambda x: str(int(x)).zfill(2))
    df_0695 = df_0695.drop('投保人证件号码', axis=1)
    info_total = pd.merge(pkl_file_df, df_0695, on='被保人身份证号_unique', how='left')
    info_total_du = info_total[info_total.duplicated(subset='车架号', keep=False)]
    if info_total_du.shape[0] > 0:
        raise ValueError('意外险匹配数据不对应')

    with open(app_path()+'/bin/data/tax_info.pkl', 'rb') as f:
        tax_dict = pickle.load(f)

    info_all = tax_dict['all_0695'][0]
    info_no_tax = tax_dict['all_0695'][1]
    info_tax = tax_dict['all_0695'][2]

    info_total['保费_0366'] = info_total['保费_0366'].apply(pd.to_numeric).round(2)
    info_total['不含税保费_0366'] = (info_total['保费_0366'] / 1.06).round(2)
    info_total['税款金额_0366'] = (info_total['保费_0366'] - info_total['不含税保费_0366']).round(2)

    info_total['不含税保费_0695'] = info_no_tax
    info_total['税额_0695'] = info_tax
    info_total['保费总计_0695'] = info_all

    info_total['车险保额（万元）'] = info_total['车险保额（万元）'].apply(pd.to_numeric).astype('int').astype('str')
    info_total['temp'] = '万'
    info_total['保额_0366'] = info_total['车险保额（万元）'] + info_total['temp']
    info_total.index = range(info_total.shape[0])
    info_total = info_total.loc[:, ['序号','投保人', '被保人', '被保人身份证号', '手机', '地区', '投保日期', '保险起期', '保险止期', '订单时间_0366', '保单号_0366', '保额_0366', '不含税保费_0366', '税款金额_0366', '保费_0366', '导入时间_0695', '保单号_0695', '不含税保费_0695', '税额_0695', '保费总计_0695', '车架号', '发动机号', '车牌号', '车辆类型', '车型名称', '车辆所属性质', '投保联系人', '代理渠道', '承保机构代码']]
    info_total = info_total.fillna('*')
    info_total = info_total.sort_values(by='车架号')
    info_total.index = range(info_total.shape[0])
    info_total.columns = range(1, info_total.shape[1] + 1)

    return info_total


def find_duplicated(file_name):
    df = pd.read_excel(file_name, dtype='str', sheet_name='承保信息汇总')
    df_du = df[df.duplicated(subset='车架号', keep=False)]
    if df_du.shape[0] > 0:
        erro_info = '原汇总表中车架号 {} 存在重复值'.format(list(df_du['车架号'].unique()))
        raise ValueError(erro_info)
    else:
        return int(df.shape[0])


def super_load_df(file_name):
    df = pd.read_excel(file_name, dtype='str', sheet_name='承保信息汇总')
    df_du = df[df.duplicated(subset='车架号', keep=False)]
    if df_du.shape[0] > 0:
        erro_info = '汇总表中车架号 {} 存在重复值'.format(list(df_du['车架号'].unique()))
        messagebox.showerror('注意!!', erro_info)
        os.rename(file_name, file_name.replace('截至20', '车架号重复-截至20'))

    return df


def add_df2old(old_total_file_name, save_total_name, info_total):
    
    cur_cn = find_duplicated(old_total_file_name)
    workbook = load_workbook(old_total_file_name)
    worksheet = workbook['承保信息汇总']
    
    info_total[1] = range(cur_cn+1, info_total.shape[0]+cur_cn+1)

    for i in range(cur_cn+2, info_total.shape[0]+cur_cn+2):
        for j in range(1, 30):
            cell_insert = worksheet.cell(row=i, column=j)
            cell_insert.font = Font(name='宋体', size=10)
            cell_insert.alignment = Alignment(horizontal='center', vertical='center')
            cell_insert.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            if j not in [13, 14, 15, 18, 19, 20]:
                cell_insert.number_format = '@'  # FORMAT_TEXT 文本格式
                cell_insert.value = str(info_total[j][i-cur_cn-2]).strip()
            else:
                cell_insert.value = float(str(info_total[j][i-cur_cn-2]).strip())

        for j in range(30, 35):
            cell_insert = worksheet.cell(row=i, column=j)
            cell_insert.font = Font(name='宋体', size=10)
            cell_insert.alignment = Alignment(horizontal='center', vertical='center')
            cell_insert.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell_insert.value = ''

    workbook.save(filename=save_total_name)
    df_complete = super_load_df(save_total_name)
    df_complete = df_complete[df_complete['投保日期'] == var_data_date.get()]
    df_complete['三责险-保费总计'] = df_complete['三责险-保费总计'].apply(pd.to_numeric).round(2)
    df_complete['驾意险-保费总计'] = df_complete['驾意险-保费总计'].apply(pd.to_numeric).round(2)
    messagebox.showinfo('{}承保信息'.format(var_data_date.get()), '承保0366共 {} 单，保费共计 {:.2f} 元；\n承保0695共 {} 单，保费共计 {:.2f} 元。'.format(df_complete.shape[0], df_complete['三责险-保费总计'].sum(), df_complete.shape[0], df_complete['驾意险-保费总计'].sum()))
    with open(app_path()+'/summary_log.txt', 'a', encoding='utf-8') as f:
        f.write('\n{} 承保三责险 {} 单，保费共计 {:.2f} 元; 承保驾意险 {} 单，保费共计 {:.2f} 元。'.format(var_data_date.get(), df_complete.shape[0], df_complete['三责险-保费总计'].sum(), df_complete.shape[0], df_complete['驾意险-保费总计'].sum()))


def create_total_sheet_main():
    try:
        file_name_0695 = app_path() + '/' + var_file_name_0695.get()
        pkl_file_name = app_path() + '/bin/data/' + var_pkl_file_name.get()
        data_date = var_data_date.get()
        save_total_name = app_path() + '/汇总表/' + '截至{}承保信息.xlsx'.format(data_date.replace('-', ''))
        old_total_file_name = app_path() + '/汇总表/' + var_old_total_file_name.get()
        info_total = create_total_sheet_raw_df(file_name_0695, pkl_file_name, data_date)
        add_df2old(old_total_file_name, save_total_name, info_total)
        info = '保存路径: ' + save_total_name
        messagebox.showinfo('处理成功!', info)
        os.remove(pkl_file_name)
    except Exception as e:
        info = '错误原因：' + str(e)
        messagebox.showerror('Error', info)


def create_total_sheet_gui():
    global window_total_sheet, var_file_name_0695, var_old_total_file_name, var_pkl_file_name, var_data_date

    window_total_sheet = Tk()
    window_total_sheet.title('摩托车汇总表生成')
    window_total_sheet.resizable(0, 0)  # 设置窗口宽高固定
    window_total_sheet.iconbitmap(app_path() + '/bin/aic_logo.ico')
    screenWidth = window_total_sheet.winfo_screenwidth()  # 获取显示区域的宽度
    screenHeight = window_total_sheet.winfo_screenheight()  # 获取显示区域的高度
    width = 560  # 设定窗口宽度
    height = 215  # 设定窗口高度
    left = int((screenWidth - width) / 2)
    top = int((screenHeight - height) / 2)-100
    window_total_sheet.geometry('{width}x{height}+{left}+{top}'.format(width=width, height=height, left=left, top=top))
    Label(window_total_sheet, text='0695承保成功调出表:', font=('Msyh', 12)).place(x=30, y=19)
    Label(window_total_sheet, text='前日汇总表:', font=('Msyh', 12)).place(x=30, y=55)
    Label(window_total_sheet, text='缓存pickle文件:', font=('Msyh', 12)).place(x=30, y=91)
    Label(window_total_sheet, text='投保日期:', font=('Msyh', 12)).place(x=30, y=127)

    with open(app_path()+'/bin/data/policy_number.pkl', 'rb') as f:
        policy_number_dict = pickle.load(f)

    # 0695承保成功调出表
    var_file_name_0695 = StringVar()
    comboxlist = ttk.Combobox(window_total_sheet, width=30, textvariable=var_file_name_0695, font=('Msyh', 12))
    list_value = [item for item in os.listdir(app_path()) if item.split('.')[-1] == 'xls' and item.split('-')[0] == policy_number_dict['all_0695']]
    list_value.sort(reverse=True)
    list_value.insert(0, '------------None------------')
    comboxlist["values"] = list_value
    try:
        comboxlist.current(1)
    except:
        comboxlist.current(0)

    comboxlist.place(x=255,y=20)


    # 前日汇总表
    var_old_total_file_name = StringVar()
    comboxlist_2 = ttk.Combobox(window_total_sheet, width=30, textvariable=var_old_total_file_name, font=('Msyh', 12))
    list_value_2 = [item for item in os.listdir(app_path() + '/汇总表/') if item.split('.')[-1] == 'xlsx']
    list_value_2.sort(reverse=True)
    comboxlist_2["values"] = list_value_2
    try:
        comboxlist_2.current(0)
    except:
        pass

    comboxlist_2.place(x=255,y=56)

    # 缓存pickle文件
    var_pkl_file_name = StringVar()
    comboxlist_3 = ttk.Combobox(window_total_sheet, width=30, textvariable=var_pkl_file_name, font=('Msyh', 12))
    list_value_3 = [item for item in os.listdir(app_path() + '/bin/data/') if item.split('.')[-1] == 'pkl' and item.split('_')[0] == 'temp']
    list_value_3.sort(reverse=True)
    comboxlist_3["values"] = list_value_3
    try:
        comboxlist_3.current(0)
    except:
        pass

    comboxlist_3.place(x=255,y=92)

    # 投保日期
    var_data_date = StringVar()
    comboxlist_4 = ttk.Combobox(window_total_sheet, width=30, textvariable=var_data_date, font=('Msyh', 12))
    list_value_4 = [datetime.now().strftime('%Y-%m-%d'), (datetime.now()-timedelta(days=1)).strftime('%Y-%m-%d'), (datetime.now()-timedelta(days=2)).strftime('%Y-%m-%d'), (datetime.now()-timedelta(days=3)).strftime('%Y-%m-%d'), (datetime.now()-timedelta(days=4)).strftime('%Y-%m-%d'), (datetime.now()-timedelta(days=5)).strftime('%Y-%m-%d')]
    comboxlist_4["values"] = list_value_4
    comboxlist_4.current(0)
    comboxlist_4.place(x=255,y=128)

    # 开始处理 按钮
    btn_start = Button(window_total_sheet, text='开始处理', command=create_total_sheet_main)
    btn_start.place(x=245, y=165)

    # 主窗口循环显示
    window_total_sheet.mainloop()
      

if __name__ == '__main__':
    create_total_sheet_gui()